
import email
import smtplib
import imaplib
from dotenv import load_dotenv
import os
import time
import sys
root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(root_path)
from src.ollama_local.main import run,run_sec
import json
import pandas as pd
from openpyxl import load_workbook
import subprocess 
from openpyxl.utils import get_column_letter
import gradio 
from email.header import decode_header
from langchain_ollama import OllamaLLM, ChatOllama
load_dotenv()

# Gmail credentials
email_address = os.getenv('EMAIL_ADDRESS')
email_password = os.getenv('EMAIL_PASSWORD')

file_to_extract = "C:/Users/dohuu/Desktop/Ollama_Host/ollama_local/src/ollama_local/tools/text_temp/extracted_images_to_text.txt"





class email_detect:
    def __init__(self):
        self.mail = imaplib.IMAP4_SSL("imap.gmail.com")
    def login_to_email(self):
        print("Đang đăng nhập ...")
        self.mail = imaplib.IMAP4_SSL("imap.gmail.com")  # Reinitialize
        status, _ = self.mail.login(email_address, email_password)
        print(f"Trạng thái đăng nhập: {status}")
        status, _ = self.mail.select('inbox')
        print(f"Trạng thái chọn inbox: {status}")
        return status


    def download_pdf_from_email(self, subject_keyword="PROCESS DOCUMENT"):
        # Search for unread emails with specified subject
        search_query = f'(UNSEEN SUBJECT "{subject_keyword}")'
        status, data = self.mail.search(None, search_query)

        if status != "OK":
            return None

        email_ids = data[0].split()
        if not email_ids:
            return None

        for num in email_ids:
            result, data = self.mail.fetch(num, "(RFC822)")
            if result != "OK":
                continue

            raw_email = data[0][1]
            msg = email.message_from_bytes(raw_email)

            subject = msg["Subject"] or ""
            print(f"📬 Đang kiểm tra email với tiêu đề: {subject}")

            for part in msg.walk():
                filename = part.get_filename()

                if filename:
                    # Decode filename if needed
                    decoded_parts = decode_header(filename)
                    filename = ''.join([
                        (text.decode(enc) if isinstance(text, bytes) else text)
                        for text, enc in decoded_parts
                    ])

                    if filename.lower().endswith(".pdf"):
                        safe_filename = filename.replace(" ", "_").replace("/", "_")
                        filepath = os.path.join("C:/Users/dohuu/Desktop/TestPDF", safe_filename)
                        with open(filepath, "wb") as f:
                            f.write(part.get_payload(decode=True))
                        print(f"✅ Đã tải xuống: {filepath}")
                        return filepath

        return None

   

    def extracted_to_json(self,file_to_extract,json_path):
        
        with open(file_to_extract,"r",encoding="utf-8") as f:
            raw_text = f.read()
        prompt = f"""
                You are an AI assistant that reads raw text extracted from a PDF invoice.

                Your job is to extract the raw txt I provided in JSON format.
                Always include the following keys (even if null) report it as like this format:

                - invoice_number
                - Báo_giá_ngày_tháng_năm 
                - STT (chỉ ghi số)
                - Mô_tả_sản_phẩm (liệt kê: tất_cả_tên_sản_phẩm, số_lượng, đơn_giá, thành_tiền,bảo_hành,ĐVT(cái,...))
                - Ghi chú
                - fee (tìm từ khóa như VAT hoặc thuế để lấy số tiền thuế, hoặc làm phép tính với cộng thanh toán và cộng thành tiền để tìm số thuế)
                - fee_percentage (tìm từ khóa như VAT ? %, hoặc phần trăm thuế hoặc làm phép tính fee hoặc thuế đã được tìm thấy với cộng thành tiền để tìm phần trăm)
                - Cộng_thành_tiền
                - Cộng_thanh_toán
                


                Dùng những key được chỉ định bên trên. Đừng tạo mới các từ khác và đừng làm giả số liệu. Nếu 
                không tìm thấy dữ liệu hãy để null.

                Raw text:
                {raw_text}
                                                        
                Return only JSON. No explanations.
                Do not return unrelated schemas such as purchase requests or contracts. 
                If the document doesn't match the schema, return null values for all keys instead of creating your own fields.
                Output only JSON.
                """
        llm = ChatOllama(
                model = "qwen2.5:7b",
                #api_key = os.getenv("API_BASE"),
                base_url = os.getenv("API_BASE")
            )



        response = llm.invoke(prompt)
        content = response.content if hasattr(response, "content") else str(response)

        
        try:
            parsed = json.loads(content)
            with open(json_path, "w", encoding="utf-8") as jf:
                json.dump(parsed, jf, indent=2, ensure_ascii=False)
            print(f"Structured JSON saved to {json_path}")
            return parsed
        except Exception as e:
            print("Failed to parse or save JSON from Ollama response.")
            print("Raw response:\n", content)
            return None
    
    


    def ask_questions_loop(self, filepath):
        run(filepath)
        while True:
            question = input("Nhập câu hỏi về file PDF (hoặc nhập /bye để thoát): ").strip()
            if question.lower() == "/bye":
                break
            run_sec(filepath, question)

        
        #json_path = file_to_extract.replace(".txt", ".json")
        #if os.path.exists(json_path):
        #    os.remove(json_path)
        #self.extracted_to_json(file_to_extract, json_path)
        #excel_path = "C:/Users/dohuu/Desktop/TestPDF/Book1.xlsx"
        #self.json_to_excel(json_path, excel_path)
        json_path = "C:/Users/dohuu/Desktop/Ollama_Host/ollama_local/src/ollama_local/tools/text_temp/extracted_images_to_text.json"

        
        if os.path.exists(json_path):
            os.remove(json_path)

       
        self.extracted_to_json(file_to_extract, json_path)

        
        excel_path = "C:/Users/dohuu/Desktop/TestPDF/Book1.xlsx"
        self.json_to_excel(json_path, excel_path)


    def json_to_excel(self, json_path, excel_path):
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        
        date = data.get("Báo_giá_ngày_tháng_năm", "")
        items = data.get("Mô_tả_sản_phẩm", [])
        total_amount = data.get("Cộng_thành_tiền", "")
        final_amount = data.get("Cộng_thanh_toán", "")
        fee = data.get("fee")
        fee_percentage = data.get("fee_percentage")
        num = data.get("STT")
        note = data.get("Ghi_chú")
        dvt = data.get("ĐVT")
        stt = data.get("STT")
        
        


        rows = []
        for item in items:
            rows.append({
                "Ngày": date,
                "STT": stt,
                "Mô tả sản phẩm" : item.get("tất_cả_tên_sản_phẩm",""),
                "ĐVT" : item.get("ĐVT"),
                "Số lượng" : item.get("số_lượng"),
                "Đơn giá": item.get("đơn_giá"),
                "Thành tiền" : item.get("thành_tiền"),
                "Ghi chú" : item.get("bảo_hành"),
                "Ghi chú thêm" : note
                
            })
        rows.append({"Cộng thành tiền" : total_amount,
                    "VAT %" : fee_percentage,
                    "VAT" : fee,
                    "Cộng thanh toán" : final_amount,})



        df = pd.DataFrame(rows, columns=["Ngày","STT","Mô tả sản phẩm","ĐVT","Số lượng","Đơn giá","Thành tiền","Ghi chú","Cộng thành tiền","VAT %","VAT","Cộng thanh toán","Ghi chú thêm"])
        df["Đơn giá"] = pd.to_numeric(df["Đơn giá"], errors="coerce").fillna(0).astype(int)
        
        if os.path.exists(excel_path):
            book = load_workbook(excel_path)
            sheet = book["Sheet1"]
            startrow = sheet.max_row
            book.close()
            mode = 'a'
            header = False  
        else:
            startrow = 0
            mode = 'w'
            header = True  

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, header=header, startrow=startrow)

        try:
            subprocess.run(["start", excel_path], check=False, shell=True)
        except Exception as e:
            print("Không thể mở Excel tự động:", e)



    #test section

    # Run initial processing once
    def process_pdf(self, filepath):
        run(filepath)  # This is your main analysis function
        return f"PDF processed: {os.path.basename(filepath)}"

    # Answer one question at a time (used in chat)
    def answer_question(self, filepath, question):
        return run_sec(filepath, question)

    # After Q&A, export to Excel
    def finalize_to_excel(self):
        json_path = "C:/Users/dohuu/Desktop/Ollama_Host/ollama_local/src/ollama_local/tools/text_temp/extracted_images_to_text.json"
    
        # Remove existing JSON file if it exists
        if os.path.exists(json_path):
            os.remove(json_path)
        
        # Check if the text file exists before processing
        if not os.path.exists(file_to_extract):
            return f"Error: Text file not found at {file_to_extract}"
        
        # Try to extract and create JSON
        try:
            parsed_data = self.extracted_to_json(file_to_extract, json_path)
            if parsed_data is None:
                return "Error: Failed to extract data to JSON"
            
            # Check if JSON file was actually created
            if not os.path.exists(json_path):
                return f"Error: JSON file was not created at {json_path}"
            
            # Proceed with Excel export
            excel_path = "C:/Users/dohuu/Desktop/TestPDF/Book1.xlsx"
            self.json_to_excel(json_path, excel_path)
            return f"Data exported to Excel: {excel_path}"
            
        except Exception as e:
            return f"Error during processing: {str(e)}"













    #simple web test
def process_email_pipeline():
    email_detection = email_detect()
    if email_detection.login_to_email() == "OK":
        
        
        filepath = email_detection.download_pdf_from_email()
        if filepath:
            email_detection.ask_questions_loop(filepath)
            return "PDF found, processed, questions answered, and Excel updated."
        else:
            return "No matching PDF email found."
    else:
        return "Failed to login to email."
email_detection = email_detect()
global_pdf_filepath = None  # For sharing PDF across interface

def fetch_and_process_pdf():
    global global_pdf_filepath
    if email_detection.login_to_email() == "OK":
        filepath = email_detection.download_pdf_from_email()
        if filepath:
            global_pdf_filepath = filepath
            process_msg = email_detection.process_pdf(filepath)
            return f"PDF downloaded and processed: {os.path.basename(filepath)}.\n{process_msg}"
        else:
            return "No matching PDF email found."
    else:
        return "Failed to login to email."

def chat_with_pdf(question, chat_history):
    if not global_pdf_filepath:
        return "Please fetch and process a PDF first.", chat_history

    answer = email_detection.answer_question(global_pdf_filepath, question)
    chat_history.append((question, answer))
    return "", chat_history

def export_excel():
    msg = email_detection.finalize_to_excel()
    return msg

with gradio.Blocks() as demo:
    gradio.Markdown("## 📧 PDF Invoice Processor + AI Q&A + Excel Export")

    fetch_button = gradio.Button("📥 Fetch and Process PDF")
    status_box = gradio.Textbox(label="System Status")

    fetch_button.click(fetch_and_process_pdf, outputs=status_box)

    chatbot = gradio.Chatbot(label="AI Q&A about PDF")
    question_input = gradio.Textbox(label="Ask a Question about the PDF")

    question_input.submit(chat_with_pdf, inputs=[question_input, chatbot], outputs=[question_input, chatbot])

    export_button = gradio.Button("📊 Export Data to Excel")
    export_result = gradio.Textbox(label="Export Status")

    export_button.click(export_excel, outputs=export_result)

demo.launch()



#if __name__ == "__main__":
    # email_detection = email_detect()
    # if email_detection.login_to_email() == "OK":
    #     print("Đăng nhập thành công! Bắt đầu quét email...\n")

    #     try:
    #         while True:
    #             print("Đang kiểm tra email mới...")
    #             status, _ = self.mail.select('inbox')
    #             filepath = email_detection.download_pdf_from_email()

    #             if filepath:
    #                 print("File PDF mới đã được tải về.")
    #                 email_detection.ask_questions_loop(filepath)
    #             else:
    #                 print("Không có email phù hợp. Sẽ kiểm tra lại sau 5 giây...\n")

    #             time.sleep(5)

    #     except KeyboardInterrupt:
    #         print("\nChương trình đã được dừng thủ công.")
    
